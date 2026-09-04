#!/usr/bin/env python3
"""RestaurantOS P0 Sprint Plan - Excel workbook for Jira/Linear import."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList

# Palette (warm earthy - consistent with PDF/PPT)
PRIMARY = "685f46"
PRIMARY_LIGHT = "d1c9b3"
ACCENT = "86702b"
ACCENT_LIGHT = "eeedea"
TEXT = "1d1c1a"
TEXT_MUTED = "8a8881"
BORDER_COLOR = "d1c9b3"
SUCCESS = "3c7a50"
WARNING = "a98846"
ERROR = "9b4a43"
INFO = "426990"
WHITE = "FFFFFF"
ROW_STRIPE = "f5f4f2"

OUTPUT = '/home/z/my-project/download/RestaurantOS-P0-Sprint-Plan.xlsx'

wb = Workbook()
wb.properties.creator = "Z.ai"
wb.properties.title = "RestaurantOS P0 Sprint Plan"
wb.properties.subject = "8-week sprint plan for P0 implementation"

# Remove default sheet
wb.remove(wb.active)

# ============================================================
# STYLES
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR),
)

title_font = Font(name='Calibri', size=18, bold=True, color=PRIMARY)
subtitle_font = Font(name='Calibri', size=11, italic=True, color=TEXT_MUTED)
header_font = Font(name='Calibri', size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type='solid')
data_font = Font(name='Calibri', size=10, color=TEXT)
data_font_bold = Font(name='Calibri', size=10, bold=True, color=TEXT)
muted_font = Font(name='Calibri', size=9, color=TEXT_MUTED, italic=True)
accent_font = Font(name='Calibri', size=10, bold=True, color=ACCENT)

row_stripe_fill = PatternFill(start_color=ROW_STRIPE, end_color=ROW_STRIPE, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
accent_light_fill = PatternFill(start_color=ACCENT_LIGHT, end_color=ACCENT_LIGHT, fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)


def style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 32


def style_data_row(ws, row, col_start, col_end, stripe=False):
    fill = row_stripe_fill if stripe else white_fill
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.fill = fill
        cell.alignment = left_align if col == col_start else center_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 24


def add_title(ws, title, subtitle, col_end=8):
    ws.cell(row=2, column=2, value=title).font = title_font
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=col_end)
    ws.cell(row=3, column=2, value=subtitle).font = subtitle_font
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=col_end)
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18


# ============================================================
# SHEET 1: SPRINT OVERVIEW
# ============================================================
ws1 = wb.create_sheet("Sprint Overview")
add_title(ws1, "RestaurantOS P0 - Sprint Overview", "8-tedenski načrt implementacije (1. oktober - 30. november 2025)", col_end=9)

# Headers at row 5
headers = ["Sprint", "Datum začetka", "Datum konca", "Cilj", "Kapaciteta (FTE)", "Napor (dnevi)", "Nalog", "Status", "Napredek"]
for col, h in enumerate(headers, 2):
    ws1.cell(row=5, column=col, value=h)
style_header_row(ws1, 5, 2, 10)

sprints = [
    ("Sprint 1", "2025-10-01", "2025-10-07", "FURS cert upload UI + Stripe POS UI začetek", 1.3, 5, 6, "Načrtovano", 0),
    ("Sprint 2", "2025-10-08", "2025-10-14", "FURS test env + Stripe POS UI zaključek", 1.3, 5, 6, "Načrtovano", 0),
    ("Sprint 3", "2025-10-15", "2025-10-21", "FURS produkcija + Stripe test kartice", 1.3, 5, 5, "Načrtovano", 0),
    ("Sprint 4", "2025-10-22", "2025-10-28", "Stripe webhook produkcija + PWA push začetek", 1.3, 5, 6, "Načrtovano", 0),
    ("Sprint 5", "2025-10-29", "2025-11-04", "PWA push notifications + install prompt", 1.3, 5, 6, "Načrtovano", 0),
    ("Sprint 6", "2025-11-05", "2025-11-11", "PWA app icon set + E2E testi", 1.3, 5, 5, "Načrtovano", 0),
    ("Sprint 7", "2025-11-12", "2025-11-18", "E2E testi + dokumentacija", 1.0, 5, 4, "Načrtovano", 0),
    ("Sprint 8", "2025-11-19", "2025-11-25", "Polishing + finalni review + deploy", 1.0, 5, 4, "Načrtovano", 0),
]

for i, sprint in enumerate(sprints):
    row = 6 + i
    for col, val in enumerate(sprint, 2):
        ws1.cell(row=row, column=col, value=val)
    style_data_row(ws1, row, 2, 10, stripe=(i % 2 == 0))
    # Format dates
    ws1.cell(row=row, column=3).number_format = 'yyyy-mm-dd'
    ws1.cell(row=row, column=4).number_format = 'yyyy-mm-dd'
    # Progress as percentage
    ws1.cell(row=row, column=10).number_format = '0%'

# Totals row
total_row = 6 + len(sprints)
ws1.cell(row=total_row, column=2, value="SKUPNO").font = data_font_bold
ws1.cell(row=total_row, column=6, value=f"=SUM(F6:F{total_row-1})").font = data_font_bold
ws1.cell(row=total_row, column=7, value=f"=SUM(G6:G{total_row-1})").font = data_font_bold
ws1.cell(row=total_row, column=8, value=f"=SUM(H6:H{total_row-1})").font = data_font_bold
ws1.cell(row=total_row, column=10, value=f"=IFERROR(SUM(J6:J{total_row-1})/8, 0)").font = data_font_bold
ws1.cell(row=total_row, column=10).number_format = '0%'
for col in range(2, 11):
    ws1.cell(row=total_row, column=col).fill = accent_light_fill
    ws1.cell(row=total_row, column=col).border = thin_border
    ws1.cell(row=total_row, column=col).alignment = center_align
ws1.row_dimensions[total_row].height = 28

# Column widths
widths = [3, 12, 14, 14, 38, 14, 12, 10, 12, 12]
for i, w in enumerate(widths):
    ws1.column_dimensions[get_column_letter(i + 1)].width = w

# Conditional formatting for Status column
ws1.conditional_formatting.add(f'I6:I{total_row-1}',
    CellIsRule(operator='equal', formula=['"Zaključeno"'], fill=PatternFill(start_color=SUCCESS, end_color=SUCCESS, fill_type='solid'), font=Font(color=WHITE, bold=True)))
ws1.conditional_formatting.add(f'I6:I{total_row-1}',
    CellIsRule(operator='equal', formula=['"V teku"'], fill=PatternFill(start_color=WARNING, end_color=WARNING, fill_type='solid'), font=Font(color=WHITE, bold=True)))
ws1.conditional_formatting.add(f'I6:I{total_row-1}',
    CellIsRule(operator='equal', formula=['"Načrtovano"'], fill=PatternFill(start_color=TEXT_MUTED, end_color=TEXT_MUTED, fill_type='solid'), font=Font(color=WHITE)))

# Progress data bar
ws1.conditional_formatting.add(f'J6:J{total_row-1}',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color=ACCENT))

ws1.freeze_panes = 'C6'

# Notes
ws1.cell(row=total_row + 3, column=2, value="Opombe:").font = data_font_bold
notes = [
    "• Kapaciteta 1.3 FTE = 1 fullstack developer (full-time) + 0.3 designer (delno)",
    "• Sprint 1-6: paralelno delo na FURS in Stripe (neodvisna modula)",
    "• Sprint 4+: PWA push notifications se začnejo po Stripe koncu",
    "• Sprint 7-8: zmanjšana kapaciteta (designer končal), fokus na testiranje in deploy",
    "• Skupni napor: 42 človek-dnevov (8 tednov × 5 dni × 1.05 FTE povprečno)",
]
for i, note in enumerate(notes):
    ws1.cell(row=total_row + 4 + i, column=2, value=note).font = muted_font
    ws1.merge_cells(start_row=total_row + 4 + i, start_column=2, end_row=total_row + 4 + i, end_column=10)

print("Sheet 1: Sprint Overview - DONE")


# ============================================================
# SHEET 2: BACKLOG (vse naloge)
# ============================================================
ws2 = wb.create_sheet("Backlog")
add_title(ws2, "P0 Backlog - vse naloge", "42 nalog razporejenih v 8 sprintov z ocenami, odvisnostmi in assignee-ji", col_end=12)

headers2 = ["ID", "Sprint", "Komponenta", "Naslov naloge", "Opis", "Tip", "Napor (dnevi)", "Story Points", "Assignee", "Odvisnosti", "Prioriteta", "Status"]
for col, h in enumerate(headers2, 2):
    ws2.cell(row=5, column=col, value=h)
style_header_row(ws2, 5, 2, 13)

# Tasks data (42 tasks)
tasks = [
    # Sprint 1
    ("P0-001", "Sprint 1", "FURS", "Zod schema za cert-upload", "Definiraj uploadCertSchema v src/lib/validations/furs.ts", "Backend", 0.5, 2, "Backend Dev", "-", "P0-Kritično", "Načrtovano"),
    ("P0-002", "Sprint 1", "FURS", "API: POST /api/furs/cert-upload", "Implementiraj route handler z multipart parsing, PKCS12 validacijo, varno shranjevanje", "Backend", 1.5, 5, "Backend Dev", "P0-001", "P0-Kritično", "Načrtovano"),
    ("P0-003", "Sprint 1", "FURS", "React: FursCertUpload komponenta", "Drag-and-drop upload z geslom in environment selectorjem", "Frontend", 2.0, 5, "Frontend Dev", "P0-002", "P0-Kritično", "Načrtovano"),
    ("P0-004", "Sprint 1", "Stripe", "Setup @stripe/react-stripe-js", "Namestitev paketa, loadStripe v layout, NEXT_PUBLIC_STRIPE_PUBLISHABLE_KEY", "Frontend", 0.5, 1, "Frontend Dev", "-", "P0-Kritično", "Načrtovano"),
    ("P0-005", "Sprint 1", "Stripe", "API: POST /api/payments/stripe-intent", "PaymentIntent creation z Zod validacijo in StripeGateway", "Backend", 1.5, 5, "Backend Dev", "-", "P0-Kritično", "Načrtovano"),
    ("P0-006", "Sprint 1", "Stripe", "React: StripeCardInput začetek", "Elements wrapper + CardElement styling", "Frontend", 1.0, 3, "Frontend Dev", "P0-004, P0-005", "P0-Kritično", "Načrtovano"),

    # Sprint 2
    ("P0-007", "Sprint 2", "FURS", "FURS TEST certifikat registracija", "Stranka pridobi test certifikat prek ToZS portala, upload v sistem", "Operativno", 2.0, 3, "Tech Lead + Stranka", "P0-003", "P0-Kritično", "Načrtovano"),
    ("P0-008", "Sprint 2", "FURS", "FURS test env validacija", "End-to-end test: izdaja računa → ZOI → SOAP → EOR prejeta", "Backend", 2.0, 5, "Backend Dev", "P0-007", "P0-Kritično", "Načrtovano"),
    ("P0-009", "Sprint 2", "FURS", "Cert-status API testiranje", "Preveri cert-status endpoint z test certifikatom", "Backend", 0.5, 2, "Backend Dev", "P0-008", "P0-Visoko", "Načrtovano"),
    ("P0-010", "Sprint 2", "Stripe", "React: StripeCardInput zaključek", "confirmCardPayment integracija, error handling, onSuccess/onError callbacks", "Frontend", 2.0, 5, "Frontend Dev", "P0-006", "P0-Kritično", "Načrtovano"),
    ("P0-011", "Sprint 2", "Stripe", "POS integracija - plačilni modal", "Integriraj StripeCardInput v POS plačilni modal (5-korak → 3-korak UX)", "Frontend", 1.5, 5, "Frontend Dev", "P0-010", "P0-Kritično", "Načrtovano"),
    ("P0-012", "Sprint 2", "Stripe", "Unit testi za stripe-intent", "Jest testi za API endpoint (validacija, error scenarios)", "Backend", 0.5, 2, "Backend Dev", "P0-005", "P0-Visoko", "Načrtovano"),

    # Sprint 3
    ("P0-013", "Sprint 3", "FURS", "FURS produkcija certifikat", "Stranka pridobi produkcijski certifikat (po test validaciji)", "Operativno", 1.0, 2, "Stranka", "P0-008", "P0-Kritično", "Načrtovano"),
    ("P0-014", "Sprint 3", "FURS", "Produkcijski preklop", "FURS_ENV=production, test z resničnim računom (1 EUR test)", "Backend", 1.0, 3, "Backend Dev", "P0-013", "P0-Kritično", "Načrtovano"),
    ("P0-015", "Sprint 3", "FURS", "Backup certifikata strategija", "Backup strategija (encrypted DB storage ali filesystem z .gitignore)", "Backend", 1.0, 2, "Backend Dev", "P0-014", "P1-Visoko", "Načrtovano"),
    ("P0-016", "Sprint 3", "Stripe", "Stripe test kartice - vse scenarije", "Testiraj 6 scenarijev (4242, 3184, 9995, 0069, 0119, 0179)", "QA", 1.5, 5, "QA", "P0-011", "P0-Kritično", "Načrtovano"),
    ("P0-017", "Sprint 3", "Stripe", "Stripe webhook produkcija setup", "Registracija webhook endpoint v Stripe dashboard, STRIPE_WEBHOOK_SECRET", "Backend", 1.0, 2, "Backend Dev", "P0-012", "P0-Kritično", "Načrtovano"),

    # Sprint 4
    ("P0-018", "Sprint 4", "Stripe", "Webhook handler posodobitev", "Posodobi /api/payment-gateways/webhook z vsemi event tipi", "Backend", 1.0, 3, "Backend Dev", "P0-017", "P0-Kritično", "Načrtovano"),
    ("P0-019", "Sprint 4", "Stripe", "Stripe CLI local testing", "Local webhook testing z stripe listen --forward-to", "Backend", 0.5, 1, "Backend Dev", "P0-018", "P0-Visoko", "Načrtovano"),
    ("P0-020", "Sprint 4", "PWA", "VAPID ključi generacija", "Generiraj VAPID ključe z web-push, dodaj v .env", "Backend", 0.5, 1, "Backend Dev", "-", "P0-Kritično", "Načrtovano"),
    ("P0-021", "Sprint 4", "PWA", "Prisma model: PushSubscription", "Dodaj PushSubscription model v schema.prisma z db:push", "Backend", 0.5, 2, "Backend Dev", "P0-020", "P0-Kritično", "Načrtovano"),
    ("P0-022", "Sprint 4", "PWA", "API: POST /api/push/subscribe", "Subscription endpoint z upsert in employeeId binding", "Backend", 1.0, 3, "Backend Dev", "P0-021", "P0-Kritično", "Načrtovano"),
    ("P0-023", "Sprint 4", "PWA", "API: POST /api/push/send", "Web-push sendNotification z VAPID config", "Backend", 1.0, 3, "Backend Dev", "P0-022", "P0-Kritično", "Načrtovano"),

    # Sprint 5
    ("P0-024", "Sprint 5", "PWA", "SW push event listener", "Dodaj push + notificationclick event v sw.js (v10)", "Frontend", 1.0, 3, "Frontend Dev", "P0-023", "P0-Kritično", "Načrtovano"),
    ("P0-025", "Sprint 5", "PWA", "React: PushSubscriptionManager", "Komponenta za subscripcijo z dovoljenjem request", "Frontend", 1.5, 5, "Frontend Dev", "P0-022", "P0-Kritično", "Načrtovano"),
    ("P0-026", "Sprint 5", "PWA", "React: InstallPrompt komponenta", "Custom install prompt z beforeinstallprompt event", "Frontend", 1.5, 3, "Frontend Dev", "-", "P0-Visoko", "Načrtovano"),
    ("P0-027", "Sprint 5", "PWA", "Push notifications testiranje", "End-to-end test: subscribe → send → prejmi notification", "QA", 1.0, 3, "QA", "P0-025, P0-024", "P0-Kritično", "Načrtovano"),
    ("P0-028", "Sprint 5", "PWA", "iOS 16.4+ compatibility check", "Testiraj push na iOS Safari 16.4+, dodaj fallback če ne deluje", "QA", 0.5, 2, "QA", "P0-027", "P1-Visoko", "Načrtovano"),
    ("P0-029", "Sprint 5", "Stripe", "3-D Secure (SCA) implementacija", "PSD2 compliant 3DS za evropske kartice (requires_action handling)", "Backend", 2.0, 5, "Backend Dev", "P0-016", "P1-Visoko", "Načrtovano"),

    # Sprint 6
    ("P0-030", "Sprint 6", "PWA", "App icon set generacija", "Generiraj 8 icon sizes (72-512px) + maskable variants + apple-touch-icon", "Designer", 1.0, 2, "Designer", "-", "P0-Visoko", "Načrtovano"),
    ("P0-031", "Sprint 6", "PWA", "manifest.json posodobitev", "Dodaj maskable purpose, shortcuts, splash_screen config", "Designer", 0.5, 1, "Designer", "P0-030", "P0-Visoko", "Načrtovano"),
    ("P0-032", "Sprint 6", "PWA", "SW v10 deploy + cache invalidation", "Update CACHE_VERSION na v10, test cache cleanup na activate", "Frontend", 0.5, 1, "Frontend Dev", "P0-024", "P0-Visoko", "Načrtovano"),
    ("P0-033", "Sprint 6", "Testi", "E2E test: FURS cert upload flow", "Playwright test za celoten FURS cert upload flow", "QA", 1.0, 3, "QA", "P0-015", "P0-Visoko", "Načrtovano"),
    ("P0-034", "Sprint 6", "Testi", "E2E test: Stripe plačilni flow", "Playwright test za Stripe plačilo z mock kartico", "QA", 1.0, 3, "QA", "P0-029", "P0-Visoko", "Načrtovano"),

    # Sprint 7
    ("P0-035", "Sprint 7", "Testi", "E2E test: PWA push notifications", "Playwright test za push subscription in notification", "QA", 1.0, 3, "QA", "P0-027", "P0-Visoko", "Načrtovano"),
    ("P0-036", "Sprint 7", "Testi", "E2E test: offline mode", "Test offline order queue + sinhronizacija", "QA", 1.0, 3, "QA", "P0-032", "P0-Visoko", "Načrtovano"),
    ("P0-037", "Sprint 7", "Testi", "Acceptance criteria validacija", "Preveri vseh 30 AC (10 na komponento) in označi", "QA", 1.5, 5, "QA", "P0-033, P0-034, P0-035", "P0-Kritično", "Načrtovano"),
    ("P0-038", "Sprint 7", "Docs", "README posodobitev", "Posodobi README z novimi P0 funkcijami in env spremenljivkami", "Tech Lead", 0.5, 1, "Tech Lead", "-", "P1-Visoko", "Načrtovano"),

    # Sprint 8
    ("P0-039", "Sprint 8", "Docs", "ARCHITECTURE.md posodobitev", "Posodobi arhitekturo z FURS/Stripe/PWA moduli", "Tech Lead", 0.5, 1, "Tech Lead", "P0-038", "P1-Visoko", "Načrtovano"),
    ("P0-040", "Sprint 8", "Docs", "PRODUCTION-LAUNCH-CHECKLIST posodobitev", "Posodobi checklist z novimi P0 koraki", "Tech Lead", 0.5, 1, "Tech Lead", "P0-039", "P0-Visoko", "Načrtovano"),
    ("P0-041", "Sprint 8", "Deploy", "Produkcijski deploy + Sentry verificiranje", "Deploy v produkcijo, preveri Sentry 24h brez napak", "Tech Lead", 1.0, 3, "Tech Lead", "P0-037, P0-040", "P0-Kritično", "Načrtovano"),
    ("P0-042", "Sprint 8", "Deploy", "Stranka onboarding + trening", "Predstavitev stranki, navodila za uporabo, trening osebja", "Tech Lead", 1.0, 2, "Tech Lead", "P0-041", "P0-Visoko", "Načrtovano"),
]

for i, task in enumerate(tasks):
    row = 6 + i
    for col, val in enumerate(task, 2):
        ws2.cell(row=row, column=col, value=val)
    style_data_row(ws2, row, 2, 13, stripe=(i % 2 == 0))

# Totals
total_row2 = 6 + len(tasks)
ws2.cell(row=total_row2, column=2, value="SKUPNO").font = data_font_bold
ws2.cell(row=total_row2, column=8, value=f"=SUM(H6:H{total_row2-1})").font = data_font_bold
ws2.cell(row=total_row2, column=9, value=f"=SUM(I6:I{total_row2-1})").font = data_font_bold
ws2.cell(row=total_row2, column=4, value=f"{len(tasks)} nalog").font = data_font_bold
for col in range(2, 14):
    ws2.cell(row=total_row2, column=col).fill = accent_light_fill
    ws2.cell(row=total_row2, column=col).border = thin_border
    ws2.cell(row=total_row2, column=col).alignment = center_align
ws2.row_dimensions[total_row2].height = 28

# Column widths
widths2 = [3, 10, 10, 12, 38, 45, 12, 12, 12, 18, 16, 14, 14]
for i, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(i + 1)].width = w

# Conditional formatting for Priority column
ws2.conditional_formatting.add(f'M6:M{total_row2-1}',
    CellIsRule(operator='equal', formula=['"P0-Kritično"'], fill=PatternFill(start_color=ERROR, end_color=ERROR, fill_type='solid'), font=Font(color=WHITE, bold=True)))
ws2.conditional_formatting.add(f'M6:M{total_row2-1}',
    CellIsRule(operator='equal', formula=['"P1-Visoko"'], fill=PatternFill(start_color=WARNING, end_color=WARNING, fill_type='solid'), font=Font(color=WHITE, bold=True)))

# Conditional formatting for Status column
ws2.conditional_formatting.add(f'N6:N{total_row2-1}',
    CellIsRule(operator='equal', formula=['"Zaključeno"'], fill=PatternFill(start_color=SUCCESS, end_color=SUCCESS, fill_type='solid'), font=Font(color=WHITE, bold=True)))
ws2.conditional_formatting.add(f'N6:N{total_row2-1}',
    CellIsRule(operator='equal', formula=['"V teku"'], fill=PatternFill(start_color=INFO, end_color=INFO, fill_type='solid'), font=Font(color=WHITE)))

# Auto-filter
ws2.auto_filter.ref = f"B5:N{total_row2-1}"
ws2.freeze_panes = 'D6'

print("Sheet 2: Backlog - DONE")


# ============================================================
# SHEET 3: TEAM CAPACITY
# ============================================================
ws3 = wb.create_sheet("Team Capacity")
add_title(ws3, "Ekipa in kapaciteta", "Člani ekipe z njihovo kapaciteto po sprintih (v FTE)", col_end=11)

headers3 = ["Član ekipe", "Vloga", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "Skupno (FTE-tedni)"]
for col, h in enumerate(headers3, 2):
    ws3.cell(row=5, column=col, value=h)
style_header_row(ws3, 5, 2, 12)

team = [
    ("Janez Novak", "Senior Fullstack Developer", 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0),
    ("Ana Horvat", "Frontend Developer", 0.5, 0.5, 0.3, 0.5, 0.8, 0.3, 0.0, 0.0),
    ("Marko Kos", "Backend Developer", 0.5, 0.5, 0.7, 0.5, 0.2, 0.0, 0.0, 0.0),
    ("Eva Zupan", "Designer (0.3 FTE)", 0.3, 0.3, 0.0, 0.0, 0.3, 0.3, 0.0, 0.0),
    ("Peter Leban", "QA Engineer", 0.0, 0.0, 0.0, 0.0, 0.0, 0.5, 1.0, 0.5),
    ("Robert Pezdirc", "Tech Lead (0.2 FTE)", 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.2, 0.5),
]

for i, member in enumerate(team):
    row = 6 + i
    for col, val in enumerate(member, 2):
        ws3.cell(row=row, column=col, value=val)
    style_data_row(ws3, row, 2, 12, stripe=(i % 2 == 0))
    # FTE values as number with 2 decimals
    for col in range(4, 12):
        ws3.cell(row=row, column=col).number_format = '0.00'
    # Total formula
    ws3.cell(row=row, column=12, value=f"=SUM(D{row}:K{row})").font = data_font_bold
    ws3.cell(row=row, column=12).number_format = '0.00'

# Totals row
total_row3 = 6 + len(team)
ws3.cell(row=total_row3, column=2, value="SKUPNA KAPACITETA").font = data_font_bold
for col in range(4, 12):
    ws3.cell(row=total_row3, column=col, value=f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row3-1})").font = data_font_bold
    ws3.cell(row=total_row3, column=col).number_format = '0.00'
ws3.cell(row=total_row3, column=12, value=f"=SUM(L6:L{total_row3-1})").font = data_font_bold
ws3.cell(row=total_row3, column=12).number_format = '0.00'
for col in range(2, 13):
    ws3.cell(row=total_row3, column=col).fill = accent_light_fill
    ws3.cell(row=total_row3, column=col).border = thin_border
    ws3.cell(row=total_row3, column=col).alignment = center_align
ws3.row_dimensions[total_row3].height = 28

# Column widths
widths3 = [3, 18, 28, 8, 8, 8, 8, 8, 8, 8, 8, 14]
for i, w in enumerate(widths3):
    ws3.column_dimensions[get_column_letter(i + 1)].width = w

# Color scale for capacity
ws3.conditional_formatting.add(f'D6:K{total_row3-1}',
    ColorScaleRule(start_type='num', start_value=0, start_color=WHITE,
                   mid_type='num', mid_value=0.5, mid_color=PRIMARY_LIGHT,
                   end_type='num', end_value=1.0, end_color=PRIMARY))

ws3.freeze_panes = 'D6'

# Notes
ws3.cell(row=total_row3 + 3, column=2, value="Opombe:").font = data_font_bold
notes3 = [
    "• FTE 1.0 = polni delovni čas (5 dni/teden)",
    "• FTE 0.5 = polovični delovni čas ali deljen projekt",
    "• Skupna kapaciteta: 33.4 FTE-tednov = 167 človek-dnevov (z rezervo za annual leave)",
    "• Rezerva: ~25% nad 42 človek-dnevov napor (za nenadne odloge in bug fix)",
    "• Designer in QA sta delno allocated - prilagodljivo glede na potrebe",
]
for i, note in enumerate(notes3):
    ws3.cell(row=total_row3 + 4 + i, column=2, value=note).font = muted_font
    ws3.merge_cells(start_row=total_row3 + 4 + i, start_column=2, end_row=total_row3 + 4 + i, end_column=12)

print("Sheet 3: Team Capacity - DONE")


# ============================================================
# SHEET 4: RISK REGISTER
# ============================================================
ws4 = wb.create_sheet("Risk Register")
add_title(ws4, "Register tveganj", "Identificirana tveganja z verjetnostjo, vplivom in mitigacijskimi ukrepi", col_end=9)

headers4 = ["ID", "Tveganje", "Kategorija", "Verjetnost", "Vpliv", "Tveganje", "Mitigacija", "Lastnik", "Status"]
for col, h in enumerate(headers4, 2):
    ws4.cell(row=5, column=col, value=h)
style_header_row(ws4, 5, 2, 10)

risks = [
    ("R-01", "Stranka ne pridobi FURS certifikata pravočasno", "Operativno", "Srednja", "Visok", "Visoko", "Začni postopek takoj (2 tedna lead time), dnevni follow-up", "Tech Lead", "Odprto"),
    ("R-02", "Stripe račun zavrnjen (KYC problemi)", "Zunanje", "Nizka", "Visok", "Srednje", "Uporabi SumUp kot backup (EU friendly), pripravi dokumentacijo vnaprej", "Tech Lead", "Odprto"),
    ("R-03", "FURS testno okolje nedosegljivo", "Tehnično", "Nizka", "Srednji", "Nizko", "Implementiraj retry z exponential backoff, fallback na simulation mode", "Backend Dev", "Odprto"),
    ("R-04", "Service Worker cache konflikti (stare verzije)", "Tehnično", "Srednja", "Nizki", "Nizko", "Force update na activate event, skipWaiting() + clients.claim()", "Frontend Dev", "Odprto"),
    ("R-05", "Push notifications ne delujejo na iOS < 16.4", "Tehnično", "Visoka", "Srednji", "Visoko", "iOS 16.4+ podpira, dodaj SMS fallback za starejše naprave", "Frontend Dev", "Odprto"),
    ("R-06", "Stripe webhook signatura neveljavna", "Tehnično", "Nizka", "Visok", "Srednje", "Testiraj z Stripe CLI locally, dokumentiraj troubleshooting", "Backend Dev", "Odprto"),
    ("R-07", "FURS rotacija certifikatov (15. sep 2025)", "Zunanje", "Visoka", "Visok", "Kritično", "cert-status API že implementiran, aktiviraj monitoring + renewal reminder", "Tech Lead", "Spremljamo"),
    ("R-08", "Izguba ključnega člana ekipe (Janez)", "Kadrovsko", "Nizka", "Visok", "Srednje", "Paralelno delo (code review), dokumentacija, knowledge sharing", "Tech Lead", "Odprto"),
    ("R-09", "Produkcijski deploy povzroči downtime", "Tehnično", "Nizka", "Visok", "Srednje", "Blue-green deploy, rollback strategija, deploy v off-peak uri (22:00)", "Tech Lead", "Odprto"),
    ("R-10", "Spremembe v FURS zakonodaji med implementacijo", "Zunanje", "Nizka", "Srednji", "Nizko", "Spremljaj FURS obvestila, agilen razvoj za hitro prilagajanje", "Tech Lead", "Odprto"),
    ("R-11", "Stripe provizije previsoke za slovenski trg", "Poslovno", "Srednja", "Srednji", "Srednje", "Pogej se o boljših pogojih (volume discount), alternativa SumUp", "Tech Lead", "Odprto"),
    ("R-12", "Testiranje nezadostno (pomanjkanje časa)", "Proces", "Srednja", "Visok", "Visoko", "TDD za kritične poti, paralelno pisanje testov z razvojem", "QA", "Odprto"),
]

for i, risk in enumerate(risks):
    row = 6 + i
    for col, val in enumerate(risk, 2):
        ws4.cell(row=row, column=col, value=val)
    style_data_row(ws4, row, 2, 10, stripe=(i % 2 == 0))

# Column widths
widths4 = [3, 8, 38, 14, 14, 14, 14, 38, 14, 12]
for i, w in enumerate(widths4):
    ws4.column_dimensions[get_column_letter(i + 1)].width = w

# Conditional formatting for Risk level
ws4.conditional_formatting.add(f'G6:G{6 + len(risks) - 1}',
    CellIsRule(operator='equal', formula=['"Kritično"'], fill=PatternFill(start_color=ERROR, end_color=ERROR, fill_type='solid'), font=Font(color=WHITE, bold=True)))
ws4.conditional_formatting.add(f'G6:G{6 + len(risks) - 1}',
    CellIsRule(operator='equal', formula=['"Visoko"'], fill=PatternFill(start_color=WARNING, end_color=WARNING, fill_type='solid'), font=Font(color=WHITE, bold=True)))
ws4.conditional_formatting.add(f'G6:G{6 + len(risks) - 1}',
    CellIsRule(operator='equal', formula=['"Srednje"'], fill=PatternFill(start_color=INFO, end_color=INFO, fill_type='solid'), font=Font(color=WHITE)))
ws4.conditional_formatting.add(f'G6:G{6 + len(risks) - 1}',
    CellIsRule(operator='equal', formula=['"Nizko"'], fill=PatternFill(start_color=SUCCESS, end_color=SUCCESS, fill_type='solid'), font=Font(color=WHITE)))

ws4.auto_filter.ref = f"B5:J{6 + len(risks) - 1}"
ws4.freeze_panes = 'D6'

print("Sheet 4: Risk Register - DONE")


# ============================================================
# SHEET 5: BURNDOWN
# ============================================================
ws5 = wb.create_sheet("Burndown")
add_title(ws5, "Burndown chart - načrtovano vs aktualno", "Sledenje napredka po sprintih (v človek-dnevih)", col_end=8)

headers5 = ["Sprint", "Načrtovano (dnevi)", "Kumulativno načrtovano", "Aktualno (dnevi)", "Kumulativno aktualno", "Preostalo (načrtovano)", "Preostalo (aktualno)"]
for col, h in enumerate(headers5, 2):
    ws5.cell(row=5, column=col, value=h)
style_header_row(ws5, 5, 2, 8)

burndown = [
    ("Sprint 1", 5, 0, 0),
    ("Sprint 2", 5, 0, 0),
    ("Sprint 3", 5, 0, 0),
    ("Sprint 4", 5, 0, 0),
    ("Sprint 5", 5, 0, 0),
    ("Sprint 6", 5, 0, 0),
    ("Sprint 7", 5, 0, 0),
    ("Sprint 8", 5, 0, 0),
]

total_planned = sum(b[1] for b in burndown)
cumulative_planned = 0
cumulative_actual = 0

for i, (sprint, planned, _, actual) in enumerate(burndown):
    row = 6 + i
    cumulative_planned += planned
    cumulative_actual += actual
    remaining_planned = total_planned - cumulative_planned
    remaining_actual = total_planned - cumulative_actual

    ws5.cell(row=row, column=2, value=sprint)
    ws5.cell(row=row, column=3, value=planned)
    ws5.cell(row=row, column=4, value=cumulative_planned)
    ws5.cell(row=row, column=5, value=actual)
    ws5.cell(row=row, column=6, value=cumulative_actual)
    ws5.cell(row=row, column=7, value=remaining_planned)
    ws5.cell(row=row, column=8, value=remaining_actual)
    style_data_row(ws5, row, 2, 8, stripe=(i % 2 == 0))

# Column widths
widths5 = [3, 12, 16, 22, 16, 22, 22, 22]
for i, w in enumerate(widths5):
    ws5.column_dimensions[get_column_letter(i + 1)].width = w

ws5.freeze_panes = 'C6'

# Create burndown chart
chart = LineChart()
chart.title = "Burndown Chart (človek-dnevi)"
chart.style = 2
chart.y_axis.title = "Preostali človek-dnevi"
chart.x_axis.title = "Sprint"
chart.height = 12
chart.width = 22

# Data: columns G (remaining planned) and H (remaining actual)
data = Reference(ws5, min_col=7, min_row=5, max_col=8, max_row=6 + len(burndown) - 1)
cats = Reference(ws5, min_col=2, min_row=6, max_row=6 + len(burndown) - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws5.add_chart(chart, f"B{6 + len(burndown) + 3}")

print("Sheet 5: Burndown - DONE")


# ============================================================
# SHEET 6: CSV EXPORT (for Jira/Linear import)
# ============================================================
ws6 = wb.create_sheet("CSV Export (Jira)")
add_title(ws6, "CSV Export za Jira/Linear import", "Kopiraj vrstice 6+ in prilepi v CSV datoteko za import v Jira/Linear/Asana", col_end=8)

# Jira CSV format headers
jira_headers = ["Summary", "Description", "Issue Type", "Priority", "Assignee", "Sprint", "Story Points", "Labels"]
for col, h in enumerate(jira_headers, 2):
    ws6.cell(row=5, column=col, value=h)
style_header_row(ws6, 5, 2, 9)

# Map our tasks to Jira format
priority_map = {
    "P0-Kritično": "Highest",
    "P1-Visoko": "High",
    "P2-Srednje": "Medium",
}

issue_type_map = {
    "Backend": "Story",
    "Frontend": "Story",
    "QA": "Task",
    "Designer": "Task",
    "Tech Lead": "Task",
    "Operativno": "Task",
    "Docs": "Task",
    "Deploy": "Task",
}

for i, task in enumerate(tasks):
    row = 6 + i
    task_id, sprint, component, title, desc, task_type, days, sp, assignee, deps, priority, status = task

    jira_row = [
        f"[{task_id}] {title}",
        f"{desc}\n\nKomponenta: {component}\nNapor: {days} dni\nOdvisnosti: {deps}",
        issue_type_map.get(task_type, "Story"),
        priority_map.get(priority, "Medium"),
        assignee,
        sprint,
        sp,
        f"P0 {component}",
    ]
    for col, val in enumerate(jira_row, 2):
        ws6.cell(row=row, column=col, value=val)
    style_data_row(ws6, row, 2, 9, stripe=(i % 2 == 0))

# Column widths
widths6 = [3, 38, 50, 10, 10, 18, 12, 10, 14]
for i, w in enumerate(widths6):
    ws6.column_dimensions[get_column_letter(i + 1)].width = w

ws6.freeze_panes = 'D6'

# Notes
notes_row = 6 + len(tasks) + 3
ws6.cell(row=notes_row, column=2, value="Navodila za import:").font = data_font_bold
import_notes = [
    "1. Izberi vrstice 6 do 47 (vse naloge)",
    "2. Kopiraj (Ctrl+C)",
    "3. Odpri Notepad/TextEdit in prilepi (Ctrl+V)",
    "4. Shrani kot 'restaurantos-p0-backlog.csv' (UTF-8 encoding)",
    "5. V Jira: System → External System Import → CSV → izberi datoteko",
    "6. Mapiraj stolpce: Summary→Summary, Description→Description, Issue Type→Issue Type, itd.",
    "7. Import in ustvari vse naloge naenkrat",
]
for i, note in enumerate(import_notes):
    ws6.cell(row=notes_row + 1 + i, column=2, value=note).font = muted_font
    ws6.merge_cells(start_row=notes_row + 1 + i, start_column=2, end_row=notes_row + 1 + i, end_column=9)

print("Sheet 6: CSV Export - DONE")


# ============================================================
# SAVE
# ============================================================
wb.save(OUTPUT)
size_kb = os.path.getsize(OUTPUT) / 1024
print(f"\n=== SPRINT PLAN XLSX GENERIRAN ===")
print(f"Datoteka: {OUTPUT}")
print(f"Velikost: {size_kb:.1f} KB")
print(f"Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
